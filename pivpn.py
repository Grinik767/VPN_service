import os
import datetime
from openpyxl import load_workbook
from txt_to_img import textfile_to_image


class Pivpn:
    def __init__(self, bash, disk):
        self.bash = bash
        self.disk = disk

    def get_list_users(self, numbers=False):
        try:
            data = self.bash.exec_command('pivpn -l')
            if data[0]:
                data = data[1]
            data = [line.split()[0] for line in data.split('\n')[2:-1]]
            if numbers:
                data = [f'{line[0] + 1}. {line[1]}' for line in list(enumerate(data))]
            return True, data
        except Exception as err:
            return False, err

    def get_list_clients(self):
        try:
            self.disk.download(src_path='/VPN/Clients.xlsx', path_or_file='Clients.xlsx')
            wb = load_workbook('Clients.xlsx')
            ws = wb.active
            list_of_clients = []
            k = 2
            for client in ws['A'][1:]:
                time_s = ws[f"E{k}"].value
                if time_s.__class__.__name__ == 'datetime':
                    time_s = time_s.strftime('%d.%m.%Y')
                time_f = ws[f"F{k}"].value
                if time_f.__class__.__name__ == 'datetime':
                    time_f = time_f.strftime('%d.%m.%Y')
                list_of_clients.append([str(k - 1), client.value, ws[f"C{k}"].value, time_s, time_f])
                k += 1
            os.remove('Clients.xlsx')
            return True, list_of_clients
        except Exception as err:
            return False, err

    def add_new_client(self, name, date_s, date_f, cost, count, count_max=0, platform='Авито', phone='', qr=True,
                       no_ads=False):
        try:
            self.disk.download(src_path='/VPN/Clients.xlsx', path_or_file='Clients.xlsx')
            if count_max == 0:
                count_max = count
            wb = load_workbook('Clients.xlsx')
            ws = wb.active
            s = 0
            for i in range(1, len(ws['A']) + 2):
                if not ws[f'A{i}'].value:
                    s = i
                    break
            nickname = f"Client{s - 1}"
            ws[f'A{s}'] = name
            ws[f'B{s}'] = nickname
            ws[f'C{s}'] = platform
            ws[f'D{s}'] = phone
            if date_s != "":
                ws[f'E{s}'] = datetime.datetime.strptime(date_s, '%d.%m.%Y')
                ws[f'E{s}'].number_format = 'DD.MM.YYYY'
            else:
                ws[f'E{s}'] = date_s
            ws[f'F{s}'] = datetime.datetime.strptime(date_f, '%d.%m.%Y')
            ws[f'F{s}'].number_format = 'DD.MM.YYYY'
            ws[f'G{s}'] = cost
            ws[f'H{s}'] = count
            ws[f'I{s}'] = count_max
            wb.save('Clients.xlsx')
            self.disk.upload(path_or_file='Clients.xlsx', dst_path='/VPN/Clients.xlsx', overwrite=True)
            os.remove('Clients.xlsx')
            if no_ads:
                self.bash.reload_file('/etc/pivpn/wireguard', 'NO_ADS.conf', 'setupVars.conf')
            else:
                self.bash.reload_file('/etc/pivpn/wireguard', 'ADS.conf', 'setupVars.conf')
            num = self.get_list_users()
            if num[0]:
                num = len(num[1])
            else:
                raise KeyError
            links = []
            for i in range(1, count + 1):
                self.bash.exec_command('pivpn add', f"{nickname}_{i}")
                num += 1
                if qr:
                    data = self.bash.exec_command('pivpn -qr', f"{num}")
                    data = '\n'.join(
                        [line.split('1m')[1:][0].strip('\x1b[0m') for line in data[1].split('\n')[num + 3:-1]])
                    with open(f"{nickname}_{i}.txt", mode='wb') as f:
                        f.write(data.encode('utf-8'))
                    textfile_to_image(f"{nickname}_{i}.txt").save(f"{nickname}_{i}.jpg")
                    os.remove(f"{nickname}_{i}.txt")
                else:
                    self.bash.download_file('/home/vpn/configs', filename=f"{nickname}_{i}.conf")
                    self.disk.upload(path_or_file=f'{nickname}_{i}.conf', dst_path=f'/VPN/{nickname}_{i}.conf',
                                     overwrite=True)
                    self.disk.publish(path=f'/VPN/{nickname}_{i}.conf')
                    links.append(self.disk.get_download_link(path=f'/VPN/{nickname}_{i}.conf'))
                    os.remove(f'{nickname}_{i}.conf')
            if not qr:
                return True, links
            return True, nickname
        except Exception as err:
            return False, err

    def delete_client(self, number):
        try:
            self.disk.download(src_path='/VPN/Clients.xlsx', path_or_file='Clients.xlsx')
            wb = load_workbook('Clients.xlsx')
            ws = wb.active
            nickname = ws[f'B{number + 1}'].value
            users = self.get_list_users()[1]
            numbers_to_delete = [i + 1 for i in range(len(users)) if users[i].startswith(nickname)]
            for i in range(len(numbers_to_delete)):
                self.bash.exec_command('pivpn -r', numbers_to_delete[i] - i, 'y')
            ws[f'A{number + 1}'], ws[f'B{number + 1}'], ws[f'C{number + 1}'], ws[f'D{number + 1}'], ws[
                f'E{number + 1}'], ws[f'F{number + 1}'], ws[f'G{number + 1}'], ws[f'H{number + 1}'], ws[
                f'I{number + 1}'] = '', '', '', '', '', '', '', '', ''
            wb.save('Clients.xlsx')
            self.disk.upload(path_or_file='Clients.xlsx', dst_path='/VPN/Clients.xlsx', overwrite=True)
            os.remove('Clients.xlsx')
            return True, 'Ok'
        except Exception as err:
            return False, err
